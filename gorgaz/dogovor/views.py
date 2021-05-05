import codecs

from django.contrib.auth import login, logout
from django.db.models import Q, Sum, Count
from django.contrib.auth.forms import AuthenticationForm
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.core.paginator import Paginator
from .models import Dogovor, Payment, Notification, Worker, Order, Plan
from .forms import DogovorForm, PaymentForm, OrderForm
from datetime import datetime, timedelta, date
import xlwt
import xlrd
from xlutils.copy import copy
import openpyxl
from docxtpl import DocxTemplate
from PIL import Image, ImageDraw, ImageFont
from io import BytesIO
import qrcode
import base64
import os
from .param import *


def main(request):
    end = datetime.now().date() + timedelta(days=EXPIRED_DAYS)
    dogovor_count = Dogovor.objects.all().count()
    dogovor_active = Dogovor.objects.filter(active=True)
    dogovor_active_count = dogovor_active.count()
    cities_active = dogovor_active.values('address_city').distinct().count
    dogovor_expiring = Dogovor.objects.filter(Q(end_date__lte=end) & Q(active=True))
    cities_expiring = dogovor_expiring.values('address_city').distinct().count
    dogovor_expiring_count = dogovor_expiring.count()
    dogovor_expired = Dogovor.objects.filter(Q(end_date__lt=datetime.now().date()) & Q(active=True))
    cities_expired = dogovor_expired.values('address_city').distinct().count
    dogovor_expired_count = dogovor_expired.count()

    year_begin = datetime.today().strftime("%Y-01-01")
    payments_year = Payment.objects.filter(date__gte=year_begin)
    payments_year_count = payments_year.count()
    amount_year = payments_year.aggregate(Sum('amount'))['amount__sum']

    year = str(datetime.today().year)
    #year = '2020'

    periods = [(year + '-01-01', year + '-01-31'),
               (year + '-02-01', year + '-02-28'),
               (year + '-03-01', year + '-03-31'),
               (year + '-04-01', year + '-04-30'),
               (year + '-05-01', year + '-05-31'),
               (year + '-06-01', year + '-06-30'),
               (year + '-07-01', year + '-07-31'),
               (year + '-08-01', year + '-08-31'),
               (year + '-09-01', year + '-09-30'),
               (year + '-10-01', year + '-10-31'),
               (year + '-11-01', year + '-11-30'),
               (year + '-12-01', year + '-12-31'),
               ]

    payments_count_month = []
    payments_amount_month = []
    for month in range(12):
        pmts = Payment.objects.filter(date__range=periods[month])
        if pmts:
            payments_count_month.append(pmts.count())
            payments_amount_month.append(pmts.aggregate(Sum('amount'))['amount__sum'])
        else:
            payments_count_month.append(0)
            payments_amount_month.append(0)

    data = {
        'title': 'Dashboard',
        'count': dogovor_count,
        'active_count': dogovor_active_count,
        'expiring_count': dogovor_expiring_count,
        'expired_count': dogovor_expired_count,
        'payments': payments_year_count,
        'payments_count_month': payments_count_month,
        'payments_amount_month': payments_amount_month,
        'amount': "{:,}".format(amount_year),
        'cities_active': cities_active,
        'cities_expiring': cities_expiring,
        'cities_expired': cities_expired,
    }
    return render(request, 'dogovor/index.html', data)


def cities_stats(request):
    end = datetime.now().date() + timedelta(days=EXPIRED_DAYS)
    today = datetime.today().date()
    active = Dogovor.objects.filter(active=True).values('address_city').annotate(active=Count('address_city'))
    expiring = Dogovor.objects.filter(Q(end_date__lte=end) & Q(end_date__gte=today) & Q(active=True)).values('address_city').annotate(expiring=Count('address_city'))
    expired = Dogovor.objects.filter(Q(end_date__lt=datetime.now().date()) & Q(active=True)).values('address_city').annotate(expired=Count('address_city'))

    for item1 in active:
        for item2 in expiring:
            if item2['address_city'] == item1['address_city']:
                item1['expiring'] = item2['expiring']

    for item1 in active:
        for item3 in expired:
            if item3['address_city'] == item1['address_city']:
                item1['expired'] = item3['expired']

    data = {
        'active': active,
        'expiring': expiring,
        'expired': expired,
    }
    return render(request, 'dogovor/cities.html', data)


def dogovor_inactive(request):
    dogovor_data = Dogovor.objects.filter(active=False).order_by('-terminate_date', '-date')
    count = dogovor_data.count()
    paginator = Paginator(dogovor_data, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    data = {
        'title': 'Расторгнутые договора',
        'dogovors': page_obj,
        'count': count,
        'query': 'Расторгнутые',
        'type': 1,
    }
    return render(request, 'dogovor/inactive.html', data)


def dogovor_inactive_search(request):
    if request.method == 'POST':
        start_date = request.POST['start']
        end_date = request.POST['end']
        dogovors_data = Dogovor.objects.filter(active=False).order_by('-terminate_date', '-date')
        dogovors_data = dogovors_data.filter(terminate_date__range=(start_date, end_date))
        count = dogovors_data.count()
    else:
        count = 0
        dogovors_data = []
    data = {
        'title': 'Расторгнутые договора',
        'dogovors': dogovors_data,
        'count': count,
        'query': 'Расторгнутые по дате',
    }
    return render(request, 'dogovor/inactivesearch.html', data)


def dogovor_expired(request):
    dogovor_data = Dogovor.objects.filter(Q(end_date__lt=datetime.now().date()) & Q(active=True)).order_by('address_city', 'address_street', 'address_house', 'address_kv')
    count = dogovor_data.count()
    paginator = Paginator(dogovor_data, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    data = {
        'title': 'Договора с просроченным сроком',
        'dogovors': page_obj,
        'count': count,
        'query': 'Просроченные',
    }
    return render(request, 'dogovor/inactive.html', data)


def user_login(request):
    if request.method == 'POST':
        form = AuthenticationForm(data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            return redirect('main')
    else:
        form = AuthenticationForm()
    data = {
        'form': form,
    }
    return render(request, 'dogovor/login.html', data)


def user_logout(request):
    logout(request)
    return redirect('login')


def dogovor_view(request, dogovor_id):
    dogovor = get_object_or_404(Dogovor, pk=dogovor_id)
    payments = Payment.objects.filter(dogovor_id=dogovor_id).order_by('date')
    orders = Order.objects.filter(dogovor_id=dogovor_id).order_by('date')
    data = {
        'dogovor': dogovor,
        'payments': payments,
        'orders': orders,
    }
    return render(request, 'dogovor/dogovor.html', data)


def dogovor_add(request):
    if request.method == 'POST':
        form = DogovorForm(request.POST)
        if form.is_valid():
            dogovor = form.save()
            return redirect('dogovor', dogovor_id=dogovor.id)
    else:
        form = DogovorForm()
    data = {
        'title': 'Новый договор',
        'form': form,
    }
    return render(request, 'dogovor/add.html', data)


def dogovor_update(request, dogovor_id):
    dogovor = Dogovor.objects.get(pk=dogovor_id)
    if request.method == 'POST':
        form = DogovorForm(request.POST, instance=dogovor)
        if form.is_valid():
            form.save()
            return redirect('dogovor', dogovor_id=dogovor.id)
    form = DogovorForm(instance=dogovor)
    data = {
        'form': form,
        'dogovor_id': dogovor_id,
    }
    return render(request, 'dogovor/update.html', data)


def dogovor_search(request):
    if request.method == 'POST':
        query = request.POST['query'].strip()
        # dogovor_data = Dogovor.objects.filter(Q(name__contains=query) | Q(number__contains=query) |
        #                                       Q(tel1__contains=query) | Q(tel2__contains=query) |
        #                                       Q(tel3__contains=query)).order_by('-date')
        #
        dogovor_data = Dogovor.objects.filter(number=query)
    else:
        dogovor_data = []
        query = ''

    data = {
        'title': 'Результат поиска',
        'dogovors': dogovor_data,
        'query': query,
    }
    return render(request, 'dogovor/search.html', data)


def dogovor_search_name(request):
    address_city = Dogovor.objects.values('address_city').distinct().order_by('address_city')
    if request.method == 'POST':
        name = request.POST['name'].strip()
        city = request.POST['address_city']
        dogovor_data = Dogovor.objects.filter(name__contains=name).order_by('address_city', 'address_street', 'address_house')
        if city:
            dogovor_data = dogovor_data.filter(address_city=city)
    else:
        dogovor_data = []
        name = ''
        city = ''

    data = {
        'title': 'Поиск по фамилии',
        'dogovors': dogovor_data,
        'cities': address_city,
        'name': name,
        'city': city,
    }
    return render(request, 'dogovor/name.html', data)


def dogovor_search_address(request):
    """Поиск договоров по адресу: населенный пункт и/или улица

    """
    exp = 1
    ul = 1
    end1 = datetime.now().date()
    end2 = datetime.now().date() + timedelta(days=EXPIRED_DAYS)
    address_city = Dogovor.objects.values('address_city').distinct().order_by('address_city')
    address_street = Dogovor.objects.values('address_street').distinct().order_by('address_street')
    if request.method == 'POST':
        city = request.POST['address_city']
        street = request.POST['address_street']
        house = request.POST['address_house']
        kv = request.POST['address_kv']
        exp = int(request.POST.get('exp'))
        ul = int(request.POST.get('ul'))
        error_message = ''
        if city and street:
            dogovor_data = Dogovor.objects.filter(Q(address_city=city) & Q(address_street=street) & Q(active=True)).order_by('address_street', 'address_house', 'address_kv')
            if house:
                dogovor_data = dogovor_data.filter(address_house=house)
            if kv:
                dogovor_data = dogovor_data.filter(address_kv=kv)
            if exp == 1:
                dogovor_data = dogovor_data.filter(end_date__lte=end1)
            if exp == 2:
                dogovor_data = dogovor_data.filter(end_date__lte=end2)
            if ul == 1:
                dogovor_data = dogovor_data.filter(fiz=True)
            if ul == 2:
                dogovor_data = dogovor_data.filter(fiz=False)
        elif city:
            dogovor_data = Dogovor.objects.filter(Q(address_city=city) & Q(active=True)).order_by('address_street', 'address_house', 'address_kv')
            if house:
                dogovor_data = dogovor_data.filter(address_house=house)
            if kv:
                dogovor_data = dogovor_data.filter(address_kv=kv)
            if exp == 1:
                dogovor_data = dogovor_data.filter(end_date__lte=end1)
            if exp == 2:
                dogovor_data = dogovor_data.filter(end_date__lte=end2)
            if ul == 1:
                dogovor_data = dogovor_data.filter(fiz=True)
            if ul == 2:
                dogovor_data = dogovor_data.filter(fiz=False)
        elif street:
            dogovor_data = []
            error_message = 'Выберите населенный пункт'
        else:
            dogovor_data = []
            error_message = 'Задайте параметры поиска'
    else:
        dogovor_data = []
        city = ''
        street = ''
        house =''
        kv = ''
        error_message = 'Параметры поиска'
    data = {
        'title': 'Поиск по адресу',
        'cities': address_city,
        'streets': address_street,
        'city': city,
        'street': street,
        'house': house,
        'kv': kv,
        'dogovors': dogovor_data,
        'message': error_message,
        'exp': exp,
        'ul': ul,
    }
    return render(request, 'dogovor/search_address.html', data)


def city_autocomplete(request):
    q = request.GET['term']
    qs = Dogovor.objects.values('address_city').distinct().order_by('address_city')
    data = []
    if q:
        qs = qs.filter(address_city__istartswith=q)
        for item in qs:
            data.append(item['address_city'])
    return JsonResponse(data, safe=False)


def street_autocomplete(request):
    q = request.GET['term']
    qs = Dogovor.objects.values('address_street').distinct().order_by('address_street')
    data = []
    if q:
        qs = qs.filter(address_street__istartswith=q)
        for item in qs:
            data.append(item['address_street'])
    return JsonResponse(data, safe=False)


def name_autocomplete(request):
    q = request.GET['term']
    qs = Dogovor.objects.values('name').distinct().order_by('name')
    data = []
    if q:
        qs = qs.filter(name__istartswith=q)
        for item in qs:
            data.append(item['name'])
    return JsonResponse(data, safe=False)


def dogovor_newpay(request, dogovor_id):
    qs = Dogovor.objects.get(pk=dogovor_id)
    workers = Worker.objects.all()
    if request.method == 'POST':
        form = PaymentForm(request.POST)
        if form.is_valid():
            Payment.objects.create(**form.cleaned_data)
            if qs.end_date is not None:
                qs.end_date = date(qs.end_date.year + 1, qs.end_date.month, qs.end_date.day)
                qs.save()
            else:
                qs.end_date = date(qs.date.year + 1, qs.date.month, qs.date.day)
                qs.save()
            return redirect('dogovor', dogovor_id=dogovor_id)
        else:
            print('invalid form')
    else:
        form = PaymentForm()
    data = {
        'title': 'Новый платеж',
        'form': form,
        'dogovor': qs,
        'workers': workers,
    }
    return render(request, 'dogovor/newpay.html', data)


def dogovor_updatepay(request, payment_id):
    payment = get_object_or_404(Payment, pk=payment_id)
    dogovor = payment.dogovor_id
    workers = Worker.objects.all()
    if request.method == 'POST':
        form = PaymentForm(request.POST, instance=payment)
        if form.is_valid():
            form.save()
            return redirect('dogovor', dogovor_id=dogovor.id)
        else:
            print('form error', form)
    form = PaymentForm(instance=payment)
    data = {
        'form': form,
        'dogovor': dogovor,
        'payment_id': payment_id,
        'workers': workers,
    }
    return render(request, 'dogovor/updatepay.html', data)


def payment_delete(request, payment_id):
    instance = get_object_or_404(Payment, pk=payment_id)
    dogovor_id = instance.dogovor_id.id
    instance.delete()
    return redirect('dogovor', dogovor_id=dogovor_id)


def payments(request):
    payments_data = Payment.objects.all().order_by('-date')[:100]
    summa = payments_data.aggregate(Sum('amount'))['amount__sum']
    data = {
        'title': 'Последние платежи',
        'payments': payments_data,
        'summa': summa,
    }
    return render(request, 'dogovor/payments.html', data)


def dogovors(request):
    dogovors_data = Dogovor.objects.all().order_by('-date')[:100]
    data = {
        'title': 'Последние договора',
        'dogovors': dogovors_data,
    }
    return render(request, 'dogovor/dogovors.html', data)


def orders(request):
    orders_data = Order.objects.all().order_by('-date', '-pk')
    paginator = Paginator(orders_data, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    data = {
        'title': 'Наряды на работы',
        'orders': page_obj,
    }
    return render(request, 'dogovor/orders.html', data)


def order_update(request, order_id):
    order = get_object_or_404(Order, pk=order_id)
    workers = Worker.objects.all()
    if request.method == 'POST':
        form = OrderForm(request.POST, instance=order)
        if form.is_valid():
            form.save()
            return redirect('plansystem')
    form = OrderForm(instance=order)
    data = {
        'form': form,
        'order': order,
        'workers': workers,
    }

    return render(request, 'dogovor/order.html', data)


def order_add(request):
    workers = Worker.objects.all()
    dogovor_id = request.GET.get('d')
    if dogovor_id:
        dogovor = Dogovor.objects.get(pk=dogovor_id)
    else:
        dogovor = None
    if request.method == 'POST':
        form = OrderForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('plansystem')
    else:
        form = OrderForm()
    data = {
        'title': 'Новый наряд',
        'form': form,
        'dogovor': dogovor,
        'workers': workers,
    }
    return render(request, 'dogovor/neworder.html', data)


def order_delete(request, order_id):
    instance = get_object_or_404(Order, pk=order_id)
    instance.delete()
    return redirect('plansystem')


def order_print(request, order_id):
    order = get_object_or_404(Order, pk=order_id)
    im = Image.open("static/images/order.png")
    blank = ImageDraw.Draw(im)
    fnt1 = ImageFont.truetype("static/font/arial.ttf", 16)
    fnt2 = ImageFont.truetype("static/font/arial.ttf", 20)
    blank.text((240, 33), str(order.pk), font=fnt1, fill=255)
    blank.text((170, 68), str(order.name), font=fnt2, fill=255)
    blank.text((130, 92), str(order.address), font=fnt2, fill=255)
    blank.text((150, 152), str(order.tel), font=fnt2, fill=255)
    if order.worker:
        blank.text((200, 234), str(order.worker), font=fnt2, fill=255)
    else:
        blank.text((200, 234), '-', font=fnt2, fill=255)
    blank.text((730, 118), str(order.date.strftime("%d.%m.%Y")), font=fnt2, fill=255)
    job = order.job
    job = job.replace(',', '\n')
    job = job.replace(';', '\n')
    if order.job:
        blank.text((55, 192), str(job), font=fnt2, fill=255)
    else:
        blank.text((55, 192), '-', font=fnt2, fill=255)

    output = BytesIO()
    im.save(output, "PNG")
    image = output.getvalue()
    contents = base64.b64encode(image).decode()
    output.close()
    data = {
        'order': order,
        'contents': contents,
    }
    return render(request, 'dogovor/printorder.html', data)


def payments_by_date(request):
    if request.method == 'POST':
        start_date = request.POST['start_date']
        end_date = request.POST['end_date']
        if end_date:
            payments_data = Payment.objects.filter(date__range=(start_date, end_date)).order_by('-date')
        else:
            payments_data = Payment.objects.filter(date__gte=start_date).order_by('-date')
        summa = payments_data.aggregate(Sum('amount'))['amount__sum']
    else:
        payments_data = []
        start_date = ''
        end_date = ''
        summa = 0
    data = {
        'title': 'Результат поиска',
        'payments': payments_data,
        'start_date': start_date,
        'end_date': end_date,
        'summa': summa,
    }
    return render(request, 'dogovor/datepayments.html', data)


def payments_by_name(request):
    if request.method == 'POST':
        name = request.POST['name'].strip()
        payments_data = Payment.objects.filter(dogovor_id__name__contains=name).order_by('-date')
        summa = payments_data.aggregate(Sum('amount'))['amount__sum']
    else:
        payments_data = []
        name = ''
        summa = 0
    data = {
        'title': 'Результат поиска',
        'payments': payments_data,
        'name': name,
        'summa': summa,
    }
    return render(request, 'dogovor/namepayments.html', data)


def payments_by_number(request):
    if request.method == 'POST':
        number = request.POST['number'].strip()
        payments_data = Payment.objects.filter(dogovor_id__number__contains=number).order_by('-date')
        summa = payments_data.aggregate(Sum('amount'))['amount__sum']
    else:
        payments_data = []
        number = ''
        summa = 0
    data = {
        'title': 'Результат поиска',
        'payments': payments_data,
        'number': number,
        'summa': summa,
    }
    return render(request, 'dogovor/numpayments.html', data)


def notifications(request):
    today = datetime.today().date()
    notification_data = Notification.objects.filter(create_time__date=today)
    data = {
        'title': 'Номера телефонов для уведомлений',
        'notifications': notification_data,
    }
    return render(request, 'dogovor/notifications.html', data)


def add_notifications(request):
    if request.method == 'POST':
        dogovors = request.POST.getlist('dogovor_id[]')
        for item in dogovors:
            dogovor = Dogovor.objects.get(id=item)
            Notification.objects.create(dogovor_id=dogovor, notify_type='Call')
    return redirect('notifications')


def update_notify1(request):
    if request.method == 'GET':
        n = request.GET['n']
        action = request.GET['action']
        notification = Notification.objects.get(pk=n)
        if action != '2':
            notification.send_time_1 = datetime.now()
            notification.success_1 = action
        else:
            notification.send_time_1 = None
            notification.success_1 = False
        notification.save()
    return redirect('notifications')


def update_notify2(request):
    if request.method == 'GET':
        n = request.GET['n']
        action = request.GET['action']
        notification = Notification.objects.get(pk=n)
        if action != '2':
            notification.send_time_2 = datetime.now()
            notification.success_2 = action
        else:
            notification.send_time_2 = None
            notification.success_2 = False
        notification.save()
    return redirect('notifications')


def dogovor_export_excel(request):
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment;filename=dogovor.xls'
    work_book = xlwt.Workbook(encoding='utf-8')
    work_sheet = work_book.add_sheet(u'Список договоров')
    borders = xlwt.Borders()
    borders.left = xlwt.Borders.THIN
    borders.right = xlwt.Borders.THIN
    borders.top = xlwt.Borders.THIN
    borders.bottom = xlwt.Borders.THIN
    style_data_row = xlwt.XFStyle()
    style_data_row.num_format_str = 'DD.MM.YYYY'
    style_data_row.borders = borders

    if request.method == 'POST':
        row = 0
        dogovors = request.POST.getlist('dogovor_id[]')
        for item in dogovors:
            dogovor = Dogovor.objects.get(id=item)
            dog = str(dogovor.number) + ' от ' + dogovor.date.strftime("%d.%m.%Y")
            work_sheet.write(row, 0, dogovor.name, style_data_row)
            work_sheet.write(row, 1, dog, style_data_row)
            work_sheet.write(row, 2, dogovor.end_date, style_data_row)
            work_sheet.write(row, 3, dogovor.tel1, style_data_row)
            work_sheet.write(row, 4, dogovor.address_city, style_data_row)
            work_sheet.write(row, 5, dogovor.address_street, style_data_row)
            work_sheet.write(row, 6, dogovor.address_house, style_data_row)
            work_sheet.write(row, 7, dogovor.address_kv, style_data_row)
            row = row + 1
        work_sheet.col(0).width = 15000
        work_sheet.col(1).width = 6000
        work_sheet.col(2).width = 3000
        work_sheet.col(3).width = 4000
        work_sheet.col(4).width = 5000
        work_sheet.col(5).width = 7000

        work_book.save(response)
        return response


def notification_export_excel(request):
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment;filename=notifications.xls'
    work_book = xlwt.Workbook(encoding='utf-8')
    work_sheet = work_book.add_sheet(u'Уведомления')
    borders = xlwt.Borders()
    borders.left = xlwt.Borders.THIN
    borders.right = xlwt.Borders.THIN
    borders.top = xlwt.Borders.THIN
    borders.bottom = xlwt.Borders.THIN
    style_data_row = xlwt.XFStyle()
    style_data_row.num_format_str = 'DD.MM.YYYY'
    style_data_row.borders = borders

    if request.method == 'POST':
        row = 0
        notifications= request.POST.getlist('notification_id[]')
        for item in notifications:
            notify = Notification.objects.get(id=item)
            dogovor = Dogovor.objects.get(id=notify.dogovor_id.id)
            phone = ''
            if dogovor.tel1:
                phone += dogovor.tel1
            if dogovor.tel2:
                if dogovor.tel1:
                    phone += ', '
                phone += dogovor.tel2
            if dogovor.tel3:
                if dogovor.tel1 or dogovor.tel2:
                    phone += ', '
                phone += dogovor.tel3
            work_sheet.write(row, 0, phone, style_data_row)
            work_sheet.write(row, 1, dogovor.name, style_data_row)
            work_sheet.write(row, 2, dogovor.address_city, style_data_row)
            work_sheet.write(row, 3, dogovor.address_street, style_data_row)
            work_sheet.write(row, 4, dogovor.address_house, style_data_row)
            work_sheet.write(row, 5, dogovor.address_kv, style_data_row)
            work_sheet.write(row, 6, dogovor.end_date, style_data_row)
            row = row + 1
        work_sheet.col(0).width = 10000
        work_sheet.col(1).width = 15000
        work_sheet.col(2).width = 5000
        work_sheet.col(3).width = 6000
        work_sheet.col(4).width = 2000
        work_sheet.col(5).width = 2000
        work_sheet.col(6).width = 3000

        work_book.save(response)
        return response

def dogovor_doc1(request, dogovor_id):
    dogovor = get_object_or_404(Dogovor, pk=dogovor_id)
    doc = DocxTemplate('dogovor/static/doc/template1.docx')

    context = {
        'number': dogovor.number,
        'date': dogovor.date.strftime("%d.%m.%Y"),
        'name': dogovor.name,
        'address': dogovor.get_full_address2(),
        'phone': dogovor.get_full_phone(),
        'sum': dogovor.amount,
    }
    doc.render(context)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document')
    response['Content-Disposition'] = 'attachment;filename=dogovor.docx'
    doc.save(response)
    return response


def dogovor_doc2(request, dogovor_id):
    dogovor = get_object_or_404(Dogovor, pk=dogovor_id)
    doc = DocxTemplate('dogovor/static/doc/template2.docx')

    context = {
        'number': dogovor.number,
        'date': dogovor.date.strftime("%d.%m.%Y"),
        'name': dogovor.name,
        'address': dogovor.get_full_address2(),
        'phone': dogovor.get_full_phone(),
        'sum': dogovor.amount,
    }
    doc.render(context)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document')
    response['Content-Disposition'] = 'attachment;filename=dogovor.docx'
    doc.save(response)
    return response


def dogovor_doc3(request, dogovor_id):
    dogovor = get_object_or_404(Dogovor, pk=dogovor_id)
    doc = DocxTemplate('dogovor/static/doc/template3.docx')

    context = {
        'number': dogovor.number,
        'date': dogovor.date.strftime("%d.%m.%Y"),
        'name': dogovor.name,
        'address': dogovor.get_full_address2(),
        'phone': dogovor.get_full_phone(),
        'sum': dogovor.amount,
    }
    doc.render(context)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document')
    response['Content-Disposition'] = 'attachment;filename=dop.docx'
    doc.save(response)
    return response


def dogovor_doc4(request, dogovor_id):
    dogovor = get_object_or_404(Dogovor, pk=dogovor_id)
    doc = DocxTemplate('dogovor/static/doc/template4.docx')

    context = {
        'number': dogovor.number,
        'date': dogovor.date.strftime("%d.%m.%Y"),
        'name': dogovor.name,
        'address': dogovor.get_full_address2(),
        'phone': dogovor.get_full_phone(),
        'sum': dogovor.amount,
    }
    doc.render(context)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document')
    response['Content-Disposition'] = 'attachment;filename=dop.docx'
    doc.save(response)
    return response


# def dogovor_doc5(request, dogovor_id):
#     dogovor = get_object_or_404(Dogovor, pk=dogovor_id)
#     doc = DocxTemplate('dogovor/static/doc/template5.docx')
#
#     context = {
#         'name': dogovor.name,
#         'address': dogovor.get_full_address2(),
#         'sum': dogovor.amount,
#     }
#     doc.render(context)
#     response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document')
#     response['Content-Disposition'] = 'attachment;filename=kvit.docx'
#     doc.save(response)
#     return response

def dogovor_doc5(request, dogovor_id):
    dogovor = get_object_or_404(Dogovor, pk=dogovor_id)
    wb = openpyxl.load_workbook(filename='dogovor/static/doc/template6.xlsm', read_only=False, keep_vba=True)
    sheet = wb['Реестр начислений']
    sheet['A4'] = dogovor.name
    sheet['B4'] = dogovor.get_full_address2()
    sheet['D4'] = dogovor.amount
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response = HttpResponse(content=openpyxl.writer.excel.save_virtual_workbook(wb))
    response['Content-Disposition'] = 'attachment; filename=kvit.xlsm'
    return response


def dogovor_doc6(request, dogovor_id):
    dog = get_object_or_404(Dogovor, pk=dogovor_id)
    im = Image.open("dogovor/static/images/kvit_template.png")
    blank = ImageDraw.Draw(im)
    fnt1 = ImageFont.truetype("dogovor/static/font/arial.ttf", 16)
    fnt2 = ImageFont.truetype("dogovor/static/font/arial.ttf", 14)
    str1 = 'Ф.И.О: ' + dog.name
    str2 = 'Адрес: ' + dog.get_full_address2() + ';'
    str3 = 'Назначение: Оплата за АДО и тех. обслуж-е'
    str4 = 'Сумма: ' + str(dog.amount) + ' руб. 00 коп.'
    blank.text((200, 160), str1, font=fnt2, fill=255)
    blank.text((200, 176), str2, font=fnt2, fill=255)
    blank.text((200, 192), str3, font=fnt2, fill=255)
    blank.text((360, 225), str4, font=fnt1, fill=255)

    blank.text((200, 466), str1, font=fnt2, fill=255)
    blank.text((200, 482), str2, font=fnt2, fill=255)
    blank.text((200, 498), str3, font=fnt2, fill=255)
    blank.text((360, 532), str4, font=fnt1, fill=255)

    qr = qrcode.QRCode(version=1, error_correction=qrcode.constants.ERROR_CORRECT_L, box_size=2, border=1, )
    data = 'ST00012|Name=ООО "Арзамасгоргаз"|PersonalAcc=40702810517500000015|BankName=ФИЛИАЛ "ЦЕНТРАЛЬНЫЙ" БАНКА ВТБ (ПАО) Г. МОСКВА|BIC=044525411|CorrespAcc=30101810145250000411|PayeelNN=5243011758|LastName='
    data = data + dog.name
    data = data + '|Purpose=Оплата за АДО и тех. обслуж-е|РауегАddress='
    data = data + dog.get_full_address2()
    data = data + '|Sum='
    data = data + str(dog.amount*100)
    data = data + '|SomeNewReq=100'
    qr.add_data(data)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    im.paste(img, (15, 400), img)
    output = BytesIO()
    im.save(output, "PNG")
    image = output.getvalue()
    contents = base64.b64encode(image).decode()
    output.close()
    data = {
        'order': dog,
        'contents': contents,
    }
    return render(request, 'dogovor/printorder.html', data)


def plan(request):
    workers = Worker.objects.all()
    plans = Plan.objects.filter(user_id=request.user)
    data = {
        'plans': plans,
        'workers': workers,
    }
    return render(request, 'dogovor/plan.html', data)


def add_plan(request):
    if request.method == 'POST':
        dogovors = request.POST.getlist('dogovor_id[]')
        for item in dogovors:
            dogovor = Dogovor.objects.get(id=item)
            Plan.objects.create(dogovor_id=dogovor, user_id=request.user)
    return redirect('plan')


def del_plan(request, plan_id):
    instance = get_object_or_404(Plan, pk=plan_id)
    instance.delete()
    return redirect('plan')


def del_all_plans(request):
    plans = Plan.objects.filter(user_id=request.user)
    plans.delete()
    return redirect('plan')


def create_orders(request):
    if request.method == 'POST':
        plans = request.POST.getlist('plan_id[]')
        job = request.POST['job']
        dt = request.POST['date']
        wrk = request.POST['worker']
        for item in plans:
            pl = Plan.objects.get(id=item)
            wr = Worker.objects.get(id=wrk)
            Order.objects.create(dogovor_id=pl.dogovor_id, name=pl.dogovor_id.name,
                                 address=pl.dogovor_id.get_full_address2(), tel=pl.dogovor_id.get_full_phone(),
                                 job=job, date=dt, worker=wr, created_by=request.user)
    return redirect('plansystem')


def plan_system(request):
    if request.method == 'POST':
        dt = request.POST['datepicker_value']
        orders_data = Order.objects.filter(completed=False).order_by('worker__name')
        orders_data = orders_data.filter(date=dt)
        cnt = orders_data.count()
    else:
        dt = datetime.today().date().strftime("%Y-%m-%d")
        orders_data = Order.objects.filter(completed=False).order_by('worker__name')
        orders_data = orders_data.filter(date=dt)
        cnt = 0
    data = {
        'orders': orders_data,
        'date': dt,
        'count': cnt,
    }
    return render(request, 'dogovor/plansystem.html', data)


def order_printall(request):
    contents_data = []
    if request.method == 'POST':
        orders_data = request.POST.getlist('order_id[]')
        for item in orders_data:
            order = Order.objects.get(id=item)
            im = Image.open("static/images/order.png")
            blank = ImageDraw.Draw(im)
            fnt1 = ImageFont.truetype("static/font/arial.ttf", 16)
            fnt2 = ImageFont.truetype("static/font/arial.ttf", 20)
            blank.text((240, 33), str(order.pk), font=fnt1, fill=255)
            blank.text((170, 68), str(order.name), font=fnt2, fill=255)
            blank.text((130, 92), str(order.address), font=fnt2, fill=255)
            blank.text((150, 152), str(order.tel), font=fnt2, fill=255)
            if order.worker:
                blank.text((200, 234), str(order.worker), font=fnt2, fill=255)
            else:
                blank.text((200, 234), '-', font=fnt2, fill=255)
            blank.text((730, 118), str(order.date.strftime("%d.%m.%Y")), font=fnt2, fill=255)
            job = order.job
            job = job.replace(',', '\n')
            job = job.replace(';', '\n')
            if order.job:
                blank.text((55, 192), str(job), font=fnt2, fill=255)
            else:
                blank.text((55, 192), '-', font=fnt2, fill=255)
            output = BytesIO()
            im.save(output, "PNG")
            image = output.getvalue()
            contents = base64.b64encode(image).decode()
            contents_data.append(contents)
            output.close()
    data = {
        'orders': orders_data,
        'contents': contents_data,
    }
    return render(request, 'dogovor/printallorders.html', data)


def order_search(request):
    if request.method == 'POST':
        query = request.POST['query'].strip()
        order_data = Order.objects.filter(pk=query)
    else:
        order_data = []
        query = ''

    data = {
        'title': 'Результат поиска',
        'orders': order_data,
        'query': query,
    }
    return render(request, 'dogovor/searchorders.html', data)


def payments_dolg(request):
    payments_data = Payment.objects.filter(dolg=True).order_by('-date')
    payments_data = payments_data.filter(dogovor_id__active=True)
    data = {
        'payments': payments_data,
    }
    return render(request, 'dogovor/dolg.html', data)


def backups(request):
    path = 'static/backups/'
    bak_list = os.listdir(path)
    print(bak_list)
    return render(request, 'dogovor/backups.html', {'backups': bak_list})
